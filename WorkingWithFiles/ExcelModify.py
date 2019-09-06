#a simple python program to make and edit an xlsx file
import openpyxl as opxl #openpyxl library for operations with excel files
import os #os module to change directory

wb = opxl.Workbook() #.Workbook creates a excel workbook 

sh = wb['Sheet'] #this created workbook above contains a sheet by default by the name 'Sheet' . So,we are accessing it here

#The below are the made by the coder to store the desired under a name
sh['A1'] = 'Name' #Accessing cell A1 to store
sh['B1'] = 'Age'  #Accessing cell B1 to store
sh['C1'] = 'Job'  #Accessing cell C1 to store

num = int(input("Enter the no of input:")) #the no of inputs 

for i in range(2,num+2): #runs the loop to take the data from the user one by one and storing it in excel file
    print()
    sh.cell(row=i,column=1).value= input("Enter the name:") 
    sh.cell(row=i,column=2).value = int(input("Enter the age:")) 
    sh.cell(row=i,column=3).value = input("Enter the Job:")


sh2 = wb.create_sheet("SampleSheet")#makes another sheet in the workbook
sh2['A1'] = "Do"
sh2['B1'] = "U"
sh2['C1'] = "No"
sh2['D1'] = "The"
sh2['E1'] = "Way" 
#storing the value in sheet 2 ie,sh2

#changing the title of the sheets 
sh.title = "SampleSheet 1" #changes the name of the sheet sh
sh2.title = "SampleSheet 2" #changes the name of the sheet sh2


os.chdir("Files") #changing directory to store the excel file

wb.save("Data.xlsx") 
#the workbook os stored in the volatile memory in the beginning,after all the operations we should manually save it like we do in Libre or Ms Excel  

wb.close()
