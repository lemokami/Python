#opens an excel file and trverse through the contents and display it better to know the no of rows and columns
import openpyxl as op
#openpycl is a python library for managing xlsx files

#use sample file in Files/sampleExcel.xlsx
#opening the file and storing the handle in the var
book = op.load_workbook("Files\\sampleExcel.xlsx") #for windows users
#book = op.load_workbook("Files/sampleExcel.xlsx") #for linux users

names = book.sheetnames #.sheetnames returns the sheetnames as a list 

sheet = book[names[0]] #takes the first sheet


for i in range(1,44): #loop to traverse through the sheet
    print(f"{i} ",end='')
    for j in range(1,8):
        print(f"{sheet.cell(row=i,column=j).value}\t",end='') #printing the sheet in the format
    print()
